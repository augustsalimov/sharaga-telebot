from dataclasses import dataclass
from datetime import datetime
from typing import Any, List, Tuple, Union

import openpyxl
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


months = {
          'январь': '01', 
          "февраль": '02', 
          "март": '03', 
          "апрель": '04', 
          "май": '05', 
          "июнь": '06',
          "июль": '07', 
          "август": '08', 
          "сентябрь": '09', 
          "октябрь": '10', 
          "ноябрь": '11', 
          "декабрь": '12'
        }

lessons_format = ('л', 'пз', 'вл', 'с', 'лаб', 'зач', 'к', 'экз')


@dataclass
class Lesson:
    date: str
    subject_name: str
    lecturer_name: str
    lesson_format: str


def create_schedule():
    """
    Обрабатывает таблицу, убирает всё ненужное и создаёт расписание
    """
    wbook, sheet = _open_sheet()
    _prepare_sheet(sheet)
    schedule = _analize_sheet(sheet)
    wbook.close()
    return schedule

def _open_sheet() -> Tuple[Workbook, Worksheet]:
    """
    Открывает таблицу. Перед этим крайне рекомендуется удалить пустые листы в документе
    и шапку <<Утверждаю>> до самой таблицы с расписанием
    """
    try:
        wbook: Workbook = openpyxl.load_workbook('425в.xlsx')
    except FileNotFoundError:
        return
    wsheet: Worksheet = wbook.active
    return wbook, wsheet

def _prepare_sheet(sheet: Worksheet) -> None:
    """
    Убирает все объединения в таблице
    """
    merged_cells = list(sheet.merged_cells.ranges)
    for item in merged_cells:
        sheet.unmerge_cells(item.coord)
        start_merged_cell_value = sheet[item.coord[:2]].value
        for row in sheet.iter_rows(min_row=item.min_row,
                                   max_row=item.min_row,
                                   min_col=item.min_col,
                                   max_col=item.max_col):
            for cell in row:
                cell.value = start_merged_cell_value

def _analize_sheet(sheet: Worksheet) -> List[Lesson]:
    """Анализирует каждую ячейку и собирает информацию по занятиям"""
    subject_column: str = None
    current_month: str = '00'
    year = str(datetime.now().year)
    lessons = []

    for col in sheet.iter_cols(max_col=30,
                               max_row=50):
        current_day: str = '00'
        for cell in col:          
            _check_weeks_row(cell, sheet)

            column_letter = _check_discipline(cell)
            if column_letter: 
                subject_column = column_letter

            month_num = _check_months(cell)
            if month_num:
                current_month = month_num

            day = _check_days(cell)
            if day:
                if current_day > day:
                    current_month = str(int(current_month) + 1)  # Если в колонке числится например март, 
                                                                 # но следом за 31 числом в той же колонке идет первое число
                                                                 # то переводим месяц на один вперед
                    current_month = current_month if len(current_month) > 1 \
                                                  else '0' + current_month
                current_day = day

            if _typer(cell) in lessons_format:
                subject_info = _get_subject_info(cell, subject_column, sheet)
                date = '-'.join((year, current_month, current_day))
                lesson = Lesson(date, **subject_info)
            
                lessons.append(lesson)  # Тогда нет смысла в этом списке,
                                        # если на лету будем добавлять в бд

    return lessons # И здесь в принципе ничего не надо возвращать

def _check_discipline(cell: Cell) -> Union[str, None]:
    """
    Проверяет, если ли в ячейке слово дисциплина
    Если есть, возвращает букву колонки, чтобы впоследствии
    проще соотносить занятие и относящийся к нему предмет
    """
    return cell.column_letter if _typer(cell) == 'дисциплина' else None

def _check_weeks_row(cell: Cell, sheet: Worksheet) -> None:
    """
    Если находит строку с четными/нечетными неделями выкидывает её
    во избежании дальнейших проблем
    """
    if _typer(cell) in ('неделя', '>'):
        sheet.delete_rows(cell.row)

def _check_months(cell: Cell) -> Union[str, None]:
    """
    Возвращает или ничего, или номер месяца
    """
    return months.get(_typer(cell))

def _check_days(cell: Cell) -> Union[str, None]:
    """
    Проверяет есть ли число в ячейке.
    Если да, возвращает его в строковом формате (для цифр 1, 2 и до 9 добавляет в начало 0)
    """
    if isinstance(cell.value, int):
        day = str(cell.value)
        return day if len(day) > 1 else '0' + day

def _get_subject_info(cell: Cell, subject_column: str, sheet: Worksheet) -> dict:
    """
    Возвращает всю информацию по предмету: 
    название, имя преподавателя, тип занятия
    """
    subject_cell = subject_column + str(cell.row)
    lecturer_cell = subject_column + str(cell.row + 1)
    lesson_format_info_cell = cell.column_letter + str(cell.row + 1)

    subject_name = sheet[subject_cell].value
    subject_lecturer = sheet[lecturer_cell].value
    lesson_format_extra_info = sheet[lesson_format_info_cell].value
    lesson_format_extra_info = '' if not lesson_format_extra_info \
                                  else ' ({})'.format(lesson_format_extra_info)
    lesson_format = cell.value + lesson_format_extra_info
    return {
        'subject_name': subject_name,
        'lecturer_name': subject_lecturer,
        'lesson_format': lesson_format
        }

def _typer(cell: Cell) -> Any:
    """
    Приводит строку к нижнему регистру.
    Если не судьба, то возвращает то же значение, что и было
    """
    try:
        cell_value = cell.value.lower()
    except AttributeError:
        cell_value = cell.value
    finally:
        return cell_value


# schedule = create_schedule()  # Проверка работы

# for lesson in schedule:
#     print(lesson.date)
#     print(lesson.subject_name)
#     print(lesson.lesson_format)
#     print(lesson.lecturer_name)
#     print('***' * 5)
#     print('\n')
